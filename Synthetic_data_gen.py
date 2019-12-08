import openpyxl
import  array
import matplotlib.pylab as plt
import math
import random
import statistics



path = "Khanapur_Flows.xlsx"

column_name = "Discharge"

wb = openpyxl.load_workbook(path)
sheet = wb.active

discharge = array.array("f",[])

for j in range(1, sheet.max_column + 1):
    if (sheet.cell(row=1, column=j).value == column_name):

        for k in range(2, sheet.max_row + 1):
            discharge.append(sheet.cell(k, j).value)


print(len(discharge))
########################################################################################################
def monthly_mean(discharge):

    sum = [0]*12

    for i in range(12):
        count = 0
        for j in range(i,len(discharge),12):
            sum[i] += discharge[j]
            count += 1


        sum[i] = sum[i] / count
        #print(sum[i])

    return sum

#print(monthly_mean(discharge))
########################################################################################################
def standard_deviation(discharge, month_mean):
    deviation = [0]*12

    for i in range(12):
        count = 0
        for j in range(i, len(discharge), 12):

            deviation[i] += (discharge[j] - month_mean[i])*(discharge[j] - month_mean[i])
            count += 1

        deviation[i] =  math.sqrt(deviation[i] / count)

    return deviation


#print(standard_deviation(discharge, monthly_mean(discharge)))
########################################################################################################
def correlation_coefficient(discharge, month_mean, monthly_deviation):
    correlation = [0]*12
    count = 0
    for i in range(0, len(discharge), 12):

        correlation[0] += ((discharge[i] - month_mean[0]) * (discharge[i + 11] - month_mean[11]))
        count += 1

    correlation[0] = correlation[0] / count
    correlation[0] = correlation[0] / (monthly_deviation[0] * monthly_deviation[11])


    for i in range(0, 11):
        count = 0
        for j in range(i , len(discharge), 12):

            correlation[i+1] +=((discharge[j] - month_mean[i]) * (discharge[j +1] - month_mean[i+1]))
            #correlation[i + 1] += ((discharge[j] - month_mean[i])*(discharge[j+1] - month_mean[i+1]))/((monthly_deviation[i])*(monthly_deviation[i+1]))
            count += 1

        correlation[i+1] = correlation[i+1]/ count
        correlation[i+1] = correlation[i+1]/ (monthly_deviation[i]*monthly_deviation[i+1])

    return correlation

#print(correlation_coefficient(discharge,monthly_mean(discharge),standard_deviation(discharge,monthly_mean(discharge))))

###############################################################################################
def regression_coefficient(correlation, deviation):
    b = [0] * 12

    b[0] = correlation[0] * deviation[0]/deviation[11]
    for i in range(11):
        b[i+1] += correlation[i+1]* deviation[i+1]/deviation[i]

    return b

###################################################################################################
p = correlation_coefficient(discharge,monthly_mean(discharge),standard_deviation(discharge,monthly_mean(discharge)))
q = standard_deviation(discharge,monthly_mean(discharge))
r = monthly_mean(discharge)
s = regression_coefficient(p,q)
#print(regression_coefficient(p,q))

#print(random.normalvariate(0,1))

def synthetic_data(r,s,q,p, discharge, n):

    previous_data = discharge[len(discharge) - 12: len(discharge)]
    #print(previous_data)
    data = [0]*n
    for i in range(n):
        x = i % 12
        data[i] = r[x]+ s[x]*(previous_data[x-1]- r[x-1])+ random.normalvariate(0,1)*q[x]* math.sqrt(1- p[x]*p[x])
        previous_data[x] = data[i]

    return data



n = int(input("Enter no. of data points: "))
print(synthetic_data(r,s,q,p, discharge, n))
w = synthetic_data(r,s,q,p, discharge, n)
k = statistics.mean(w)
l = statistics.stdev(w,k)
print(f"""
Mean: {k}
Standard Deviation: {l}""")





